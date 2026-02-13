"""
Excel handler – reads and writes CRM data from/to .xlsx workbooks.

Each synced object gets its own worksheet.  The first row is always
a header row.  Column A is always the Twenty record ``id`` and
column B is ``updatedAt``.  The remaining columns map to the
configured field list.
"""

from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.worksheet import Worksheet

from .config import EXCEL_FILE_PATH, SYNC_OBJECTS

logger = logging.getLogger(__name__)

# ── styling constants ────────────────────────────────────────────────
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

# Reserved columns present in every sheet
_RESERVED_COLS = ["id", "updatedAt"]


def _all_columns(object_key: str) -> list[str]:
    """Return the full ordered column list for an object."""
    return _RESERVED_COLS + SYNC_OBJECTS[object_key]["fields"]


# ── workbook helpers ─────────────────────────────────────────────────


def _ensure_workbook(path: str = EXCEL_FILE_PATH) -> Workbook:
    """Open an existing workbook or create a new one."""
    p = Path(path)
    if p.exists():
        return load_workbook(p)
    wb = Workbook()
    # Remove default sheet created by openpyxl
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


def _ensure_sheet(wb: Workbook, object_key: str) -> Worksheet:
    """Return the worksheet for *object_key*, creating it if needed."""
    sheet_name = SYNC_OBJECTS[object_key]["sheet_name"]
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)
    cols = _all_columns(object_key)
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    # Auto-fit a reasonable default width
    for col_idx in range(1, len(cols) + 1):
        ws.column_dimensions[
            ws.cell(row=1, column=col_idx).column_letter
        ].width = 22
    ws.freeze_panes = "A2"
    return ws


# ── flattening / unflattening nested CRM data ───────────────────────

def _flatten_value(value: Any) -> Any:
    """
    Twenty returns composite fields as dicts, e.g.
    ``{"primaryEmail": "a@b.com"}`` or ``{"primaryLinkUrl": "…"}``.
    Flatten them to a single displayable value.
    """
    if isinstance(value, dict):
        # name composites: {"firstName": "", "lastName": ""}
        if "firstName" in value or "lastName" in value:
            first = value.get("firstName", "") or ""
            last = value.get("lastName", "") or ""
            return f"{first} {last}".strip()
        # email composites
        if "primaryEmail" in value:
            return value["primaryEmail"]
        # phone composites
        if "primaryPhoneNumber" in value:
            return value["primaryPhoneNumber"]
        # link composites (linkedin, etc.)
        if "primaryLinkUrl" in value:
            return value["primaryLinkUrl"]
        if "primaryLinkLabel" in value:
            return value["primaryLinkLabel"]
        # currency composites
        if "amountMicros" in value:
            return value["amountMicros"] / 1_000_000
        # address composites
        if "addressStreet1" in value:
            parts = [
                value.get("addressStreet1", ""),
                value.get("addressStreet2", ""),
                value.get("addressCity", ""),
                value.get("addressState", ""),
                value.get("addressPostcode", ""),
                value.get("addressCountry", ""),
            ]
            return ", ".join(p for p in parts if p)
        # fallback – return str representation
        return str(value)
    return value


def _unflatten_value(field: str, value: Any, existing: Any = None) -> Any:
    """
    Reverse of ``_flatten_value``: rebuild the composite dict expected
    by the Twenty API from a flat Excel cell value.
    """
    if existing is not None and isinstance(existing, dict):
        # Use the existing shape as a template
        if "firstName" in existing or "lastName" in existing:
            parts = str(value).split(" ", 1) if value else ["", ""]
            return {
                "firstName": parts[0],
                "lastName": parts[1] if len(parts) > 1 else "",
            }
        if "primaryEmail" in existing:
            return {"primaryEmail": str(value) if value else ""}
        if "primaryPhoneNumber" in existing:
            return {"primaryPhoneNumber": str(value) if value else ""}
        if "primaryLinkUrl" in existing:
            return {"primaryLinkUrl": str(value) if value else ""}
        if "amountMicros" in existing:
            try:
                return {
                    "amountMicros": int(float(value) * 1_000_000),
                    "currencyCode": existing.get("currencyCode", "USD"),
                }
            except (TypeError, ValueError):
                return existing
        if "addressStreet1" in existing:
            # Simple: put entire string into street1
            return {**existing, "addressStreet1": str(value) if value else ""}
        return existing  # unknown composite → keep original

    # Heuristic: if existing is None we infer shape from the field name.
    field_lower = field.lower()
    if field_lower == "name":
        parts = str(value).split(" ", 1) if value else ["", ""]
        return {"firstName": parts[0], "lastName": parts[1] if len(parts) > 1 else ""}
    if field_lower in ("email", "emails"):
        return {"primaryEmail": str(value) if value else ""}
    if field_lower in ("phone", "phones"):
        return {"primaryPhoneNumber": str(value) if value else ""}
    if "link" in field_lower:
        return {"primaryLinkUrl": str(value) if value else ""}
    if field_lower in ("address",):
        return {"addressStreet1": str(value) if value else ""}
    if field_lower in ("annualrecurringrevenue", "amount"):
        try:
            return {"amountMicros": int(float(value) * 1_000_000), "currencyCode": "USD"}
        except (TypeError, ValueError):
            return {"amountMicros": 0, "currencyCode": "USD"}
    return value


# ── public interface ─────────────────────────────────────────────────


def read_excel(
    object_key: str, path: str = EXCEL_FILE_PATH
) -> list[dict[str, Any]]:
    """
    Read all rows from the worksheet for *object_key* and return them
    as a list of dicts keyed by column name.
    """
    p = Path(path)
    if not p.exists():
        logger.info("Excel file does not exist yet – returning empty list")
        return []

    wb = load_workbook(p, data_only=True)
    sheet_name = SYNC_OBJECTS[object_key]["sheet_name"]
    if sheet_name not in wb.sheetnames:
        logger.info("Sheet '%s' not found – returning empty list", sheet_name)
        return []

    ws = wb[sheet_name]
    cols = _all_columns(object_key)
    rows: list[dict[str, Any]] = []
    for row_idx in range(2, ws.max_row + 1):
        record: dict[str, Any] = {}
        all_none = True
        for col_idx, col_name in enumerate(cols, start=1):
            val = ws.cell(row=row_idx, column=col_idx).value
            record[col_name] = val
            if val is not None:
                all_none = False
        if all_none:
            continue
        record["_excel_row"] = row_idx
        rows.append(record)
    return rows


def write_excel(
    object_key: str,
    crm_records: list[dict[str, Any]],
    path: str = EXCEL_FILE_PATH,
) -> None:
    """
    (Over)write the worksheet for *object_key* with *crm_records*.
    Existing data in the sheet is replaced.
    """
    wb = _ensure_workbook(path)
    sheet_name = SYNC_OBJECTS[object_key]["sheet_name"]

    # Remove existing sheet to rebuild
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = _ensure_sheet(wb, object_key)
    cols = _all_columns(object_key)

    for row_offset, record in enumerate(crm_records):
        row_idx = row_offset + 2  # row 1 is the header
        for col_idx, col_name in enumerate(cols, start=1):
            raw = record.get(col_name)
            ws.cell(row=row_idx, column=col_idx, value=_flatten_value(raw))

    wb.save(path)
    logger.info(
        "Wrote %d records to sheet '%s' in %s",
        len(crm_records),
        sheet_name,
        path,
    )


def upsert_excel_rows(
    object_key: str,
    records: list[dict[str, Any]],
    path: str = EXCEL_FILE_PATH,
) -> None:
    """
    Insert or update rows in the Excel sheet.  Matches on ``id``.
    """
    wb = _ensure_workbook(path)
    ws = _ensure_sheet(wb, object_key)
    cols = _all_columns(object_key)

    # Build id → row index map from existing data
    id_col = cols.index("id") + 1
    existing_ids: dict[str, int] = {}
    for row_idx in range(2, ws.max_row + 1):
        rid = ws.cell(row=row_idx, column=id_col).value
        if rid:
            existing_ids[str(rid)] = row_idx

    for record in records:
        if not isinstance(record, dict):
            logger.warning("Skipping non-dict record in upsert: %r", record)
            continue
        rid = str(record.get("id", ""))
        if rid in existing_ids:
            target_row = existing_ids[rid]
        else:
            target_row = ws.max_row + 1
            existing_ids[rid] = target_row

        for col_idx, col_name in enumerate(cols, start=1):
            raw = record.get(col_name)
            ws.cell(row=target_row, column=col_idx, value=_flatten_value(raw))

    wb.save(path)
    logger.info("Upserted %d records in '%s'", len(records), SYNC_OBJECTS[object_key]["sheet_name"])
